import sys
import json
import argparse
from datetime import datetime
from pathlib import Path
import openpyxl

TRACKER = Path(__file__).parent / "IG_Engagement_Tracker.xlsx"


def load_wb():
    return openpyxl.load_workbook(TRACKER)


def save_wb(wb):
    wb.save(TRACKER)


def write_picks(data: list[dict]):
    """Overwrite Today's Picks tab with today's 5 picks."""
    wb = load_wb()
    ws = wb["Today's Picks"]
    for row in ws.iter_rows(min_row=2, max_row=6):
        for cell in row:
            cell.value = None
    for i, pick in enumerate(data[:5], start=2):
        ws.cell(i, 1, pick.get("date", datetime.today().strftime("%Y-%m-%d")))
        ws.cell(i, 2, pick.get("bucket", ""))
        ws.cell(i, 3, pick.get("handle", ""))
        ws.cell(i, 4, pick.get("followers", ""))
        ws.cell(i, 5, pick.get("url", ""))
        ws.cell(i, 6, pick.get("caption_snippet", ""))
        ws.cell(i, 7, pick.get("why", ""))
        ws.cell(i, 8, pick.get("suggested_comment", ""))
    save_wb(wb)
    print(f"Wrote {len(data)} picks to Today's Picks tab.")


def append_log(data: list[dict]):
    """Append rows to Comment Log and back-fill Your Comment/Verdict/Replied in Today's Picks."""
    wb = load_wb()
    ws_log = wb["Comment Log"]
    ws_picks = wb["Today's Picks"]

    # Build URL and handle index for Today's Picks rows 2-6
    picks_index = {}
    for row in ws_picks.iter_rows(min_row=2, max_row=6):
        url = str(row[4].value or "").strip()
        handle = str(row[2].value or "").strip().lower().lstrip("@")
        if url:
            picks_index[url] = row
        if handle:
            picks_index[handle] = row

    next_row = ws_log.max_row + 1
    for pick in data:
        # Append to Comment Log
        ws_log.cell(next_row, 1, pick.get("date", datetime.today().strftime("%Y-%m-%d")))
        ws_log.cell(next_row, 2, pick.get("handle", ""))
        ws_log.cell(next_row, 3, pick.get("bucket", ""))
        ws_log.cell(next_row, 4, pick.get("url", ""))
        ws_log.cell(next_row, 5, pick.get("suggested_comment", ""))
        ws_log.cell(next_row, 6, pick.get("your_comment", ""))
        ws_log.cell(next_row, 7, pick.get("verdict", ""))
        ws_log.cell(next_row, 8, pick.get("replied", ""))
        ws_log.cell(next_row, 9, pick.get("followed", ""))
        ws_log.cell(next_row, 10, pick.get("notes", ""))
        next_row += 1

        # Back-fill Today's Picks — match by URL first, then handle
        url = pick.get("url", "").strip()
        handle = pick.get("handle", "").strip().lower().lstrip("@")
        picks_row = picks_index.get(url) or picks_index.get(handle)
        if picks_row:
            picks_row[8].value = pick.get("your_comment", "")   # col 9 — Your Comment
            picks_row[9].value = pick.get("verdict", "")         # col 10 — Verdict
            picks_row[10].value = pick.get("replied", "")        # col 11 — Replied?

    save_wb(wb)
    print(f"Appended {len(data)} rows to Comment Log and updated Today's Picks.")


def update_follow(handle: str, followed: str):
    """Update Followed Me? in Comment Log and Followed? in Today's Picks for a handle."""
    wb = load_wb()
    clean_handle = handle.lower().lstrip("@")

    # Update Comment Log — col 9 (Followed Me?)
    ws_log = wb["Comment Log"]
    log_updated = 0
    for row in ws_log.iter_rows(min_row=2):
        if str(row[1].value or "").lower().lstrip("@") == clean_handle:
            if not row[8].value or row[8].value == "":
                row[8].value = followed
                log_updated += 1

    # Update Today's Picks — col 12 (Followed?) if handle is still in today's tab
    ws_picks = wb["Today's Picks"]
    picks_updated = 0
    for row in ws_picks.iter_rows(min_row=2, max_row=6):
        if str(row[2].value or "").lower().lstrip("@") == clean_handle:
            row[11].value = followed  # col 12 — Followed?
            picks_updated += 1

    save_wb(wb)
    print(f"Updated follow status for {handle}: {log_updated} Comment Log rows, {picks_updated} Today's Picks rows.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="cmd")

    p = sub.add_parser("picks")
    p.add_argument("--data", required=True, help="JSON array of pick dicts")

    l = sub.add_parser("log")
    l.add_argument("--data", required=True, help="JSON array of comment dicts")

    f = sub.add_parser("follow")
    f.add_argument("--handle", required=True)
    f.add_argument("--followed", required=True, choices=["yes", "no", "unknown"])

    args = parser.parse_args()

    if args.cmd == "picks":
        write_picks(json.loads(args.data))
    elif args.cmd == "log":
        append_log(json.loads(args.data))
    elif args.cmd == "follow":
        update_follow(args.handle, args.followed)
    else:
        parser.print_help()
        sys.exit(1)
